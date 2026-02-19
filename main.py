from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from models import ServicoPosVenda

app = FastAPI()
ARQUIVO_RETORNO = "planilha_confirmada.xlsx"

ESTILO_CSS = """
<style>
    :root { --primaria: #800080; --secundaria: #ff69b4; --fundo: #f8f9fa; --whatsapp: #25D366; }
    body { font-family: 'Segoe UI', sans-serif; background-color: var(--fundo); margin: 0; padding: 0; }
    .header { background: linear-gradient(135deg, var(--primaria), var(--secundaria)); color: white; padding: 40px 20px; text-align: center; }
    .container { max-width: 800px; margin: -30px auto 50px; padding: 20px; }
    .cliente-card { 
        background: white; border-radius: 12px; padding: 20px; margin-bottom: 15px; 
        display: flex; justify-content: space-between; align-items: center;
        box-shadow: 0 2px 5px rgba(0,0,0,0.05); border-left: 5px solid var(--primaria);
    }
    .btn-whats { 
        background-color: var(--whatsapp); color: white; padding: 12px 25px; 
        text-decoration: none; border-radius: 30px; font-weight: bold;
    }
    .btn-whats:visited { background-color: #6c757d !important; }
    .btn-baixar { display: inline-block; background: var(--primaria); color: white; padding: 15px 30px; border-radius: 8px; text-decoration: none; font-weight: bold; margin-bottom: 20px; }
</style>
"""

@app.get("/", response_class=HTMLResponse)
async def principal():
    return f"""
    <html>
        <head><title>My Acessórios</title>{ESTILO_CSS}</head>
        <body>
            <div class="header"><h1>My Acessórios 🌸</h1><p>Sistema de Pós-Venda</p></div>
            <div class="container" style="text-align:center; background:white; padding:30px; border-radius:15px;">
                <form action="/gerar-links" enctype="multipart/form-data" method="post">
                    <input name="file" type="file" accept=".xlsx" required><br><br>
                    <button type="submit" style="background:var(--primaria); color:white; padding:10px 30px; border:none; border-radius:20px; cursor:pointer;">GERAR LISTA</button>
                </form>
            </div>
        </body>
    </html>
    """

@app.post("/gerar-links", response_class=HTMLResponse)
async def gerar_links(file: UploadFile = File(...)):
    conteudo = await file.read()
    df = pd.read_excel(io.BytesIO(conteudo))
    df.columns = [str(col).strip().lower() for col in df.columns]
    
    wb = load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    roxo_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

    html_links = f"<html><head>{ESTILO_CSS}</head><body><div class='header'><h1>Lista de Envios 📲</h1></div><div class='container'><div style='text-align:center'><a href='/baixar-planilha' class='btn-baixar'>📥 Baixar Planilha Marcada</a></div>"

    for index, linha in df.iterrows():
        try:
            # Pegando os dados (vendedora e cliente agora batem com o models.py)
            vendedora = str(linha.get('vendedora', 'Equipe')).strip()
            cliente = str(linha.get('cliente', 'Cliente')).strip()
            ddd = str(linha.get('ddd', '83')).replace(".0", "").strip()
            telefone = str(linha.get('telefone', '')).replace(".0", "").strip()

            if vendedora.lower() in ['nan', '', 'vendedora']: continue

            servico = ServicoPosVenda(vendedora=vendedora, cliente=cliente, ddd=ddd, telefone=telefone)
            link = servico.gerar_link_whatsapp()
            
            html_links += f"""
                <div class="cliente-card">
                    <div>
                        <strong style="font-size: 1.1em;">👤 {servico.cliente.title()}</strong><br>
                        <small style="color: #888;">Atendente: {servico.vendedora.title()}</small>
                    </div>
                    <a href="{link}" target="_blank" class="btn-whats">ENVIAR WHATSAPP</a>
                </div>
            """
            ws.cell(row=index + 2, column=7).fill = roxo_fill
        except Exception as e:
            print(f"Erro na linha {index}: {e}")

    wb.save(ARQUIVO_RETORNO)
    html_links += "</div></body></html>"
    return html_links

@app.get("/baixar-planilha")
async def baixar_planilha():
    return FileResponse(ARQUIVO_RETORNO, filename="pos_venda_concluido.xlsx")
